import re
import os
import requests
from io import BytesIO
from PIL import Image
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager

# 데이터를 엑셀에 작성하고 이미지 포함
def create_excel_with_images(url_list, file_name):
    wb = openpyxl.Workbook()
    ws = wb.active

    # 헤더 작성
    headers = ['링크', '제품명', '상세정보', '금액'] + [f'옵션 {i+1}' for i in range(15)] + ['이미지 폴더']
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    # 이미지 다운로드 및 저장 함수
    def download_image(image_url, folder_path, image_name):
        try:
            response = requests.get(image_url, timeout=10)
            image_data = response.content
            image_stream = BytesIO(image_data)
            img = Image.open(image_stream)
            if img.format == 'WEBP':
                img = img.convert('RGB')
            img.save(os.path.join(folder_path, image_name))
        except requests.exceptions.Timeout:
            print(f"Timeout downloading image {image_url}")
        except Exception as e:
            print(f"Error downloading image {image_url}: {e}")

    # 각 URL을 순회하며 데이터 추출 및 엑셀에 작성
    row_index = 2
    for url_index, url in enumerate(url_list, start=1):
        print(f"Processing URL: {url}")
        try:
            # 웹 드라이버 설정
            options = webdriver.ChromeOptions()
            options.add_argument('--headless')  # 브라우저 창을 열지 않고 실행
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')

            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            driver.set_page_load_timeout(30)

            try:
                driver.get(url)
            except TimeoutException:
                print(f"Timeout loading URL: {url}")
                driver.quit()
                continue

            # 404 페이지 확인
            try:
                error_element = driver.find_element(By.XPATH, "//*[contains(text(), '404') or contains(text(), 'Page Not Found')]")
                if error_element:
                    print(f"URL not found: {url}")
                    driver.quit()
                    continue
            except NoSuchElementException:
                pass

            wait = WebDriverWait(driver, 10)

            # 데이터 추출
            title_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'view_tit')))
            title = title_element.text
            print(f"Title: {title}")

            # 폴더 생성
            sanitized_title = re.sub(r'[\\/*?:"<>|]', '', title)  # 폴더명으로 사용할 수 없는 문자 제거
            folder_name = f"{url_index}_{sanitized_title}"
            folder_path = os.path.join(os.getcwd(), folder_name)
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)

            price_element = driver.find_element(By.CLASS_NAME, 'real_price')
            price = price_element.text
            print(f"Price: {price}")

            option_elements = driver.find_elements(By.CSS_SELECTOR, 'div.dropdown-item')
            options = []
            for option in option_elements:
                try:
                    a_element = option.find_element(By.CSS_SELECTOR, 'a._requireOption')
                    option_html = a_element.get_attribute('outerHTML')
                    size_match = re.search(r'<span class="blocked margin-bottom-lg">(.*?)</span>', option_html)
                    price_match = re.search(r'<span class="no-margin blocked"><strong>(.*?)</strong></span>', option_html)
                    size = size_match.group(1).strip() if size_match else 'N/A'
                    option_price = price_match.group(1).strip() if price_match else 'N/A'
                    options.append(f'Size: {size}, Price: {option_price}')
                except Exception as e:
                    options.append('Option not found')
            print(f"Options: {options}")

            goods_thumbs_element = driver.find_element(By.CLASS_NAME, 'goods_thumbs')
            thumb_images = goods_thumbs_element.find_elements(By.TAG_NAME, 'img')
            thumb_image_srcs = [img.get_attribute('src') for img in thumb_images]

            detail_element = driver.find_element(By.CLASS_NAME, 'detail_detail_wrap')
            def extract_text_and_image_srcs(element):
                imgs = element.find_elements(By.TAG_NAME, 'img')
                image_srcs = [img.get_attribute('src') for img in imgs]
                driver.execute_script("""
                var imgs = arguments[0].getElementsByTagName('img');
                while(imgs.length > 0) {
                    imgs[0].parentNode.removeChild(imgs[0]);
                }
                """, element)
                text = element.text
                cleaned_text = ' '.join(text.split())
                return cleaned_text, image_srcs

            detail_text, detail_image_srcs = extract_text_and_image_srcs(detail_element)
            print(f"Detail Text: {detail_text}")

            # 엑셀에 데이터 작성
            ws.cell(row=row_index, column=1, value=url)
            ws.cell(row=row_index, column=2, value=title)
            ws.cell(row=row_index, column=3, value=detail_text)
            ws.cell(row=row_index, column=4, value=price)
            for i, option in enumerate(options, start=5):
                ws.cell(row=row_index, column=i, value=option)
            ws.cell(row=row_index, column=20, value=folder_name)  # 폴더 이름을 기록

            # 이미지 다운로드
            for idx, image_url in enumerate(thumb_image_srcs):
                download_image(image_url, folder_path, f'product_image_{idx+1}.png')

            for idx, image_url in enumerate(detail_image_srcs):
                download_image(image_url, folder_path, f'detail_image_{idx+1}.png')

            row_index += 1

            driver.quit()

        except WebDriverException as e:
            print(f"WebDriverException processing URL {url}: {e}")
        except Exception as e:
            print(f"Error processing URL {url}: {e}")

    # 엑셀 파일 저장
    wb.save(file_name)
    print(f"Excel file '{file_name}' created successfully.")

# URL 리스트 생성
url_list = [f'https://replview.com/replicaclothing/?idx={i}' for i in range(1, 10001)]

# 엑셀 파일 생성
file_name = '상품 정보.xlsx'
create_excel_with_images(url_list, file_name)
